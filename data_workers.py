from collections import defaultdict

import openpyxl
import pandas as pd

from classes import *


def parse_groups_subjects(pi_text):
    subjects = []
    subject_parts = pi_text.split("-")

    for part in subject_parts:
        if "(" in part and ")" in part:
            subject_name = part.split("(")[0].strip()
            details_text = part.split("(")[1].rstrip().rstrip(")")
            details = []
            for detail in details_text.split(","):
                parts = detail.split("|")
                type_ = parts[0].strip()
                hours = float(parts[1])
                subgroups = int(parts[2]) if len(parts) > 2 else None
                details.append(SubjectDetail(type_, hours, subgroups))
            subjects.append(Subject(subject_name, details))
    return subjects


def parse_teachers_subjects(text):
    subjects = []
    subject_parts = text.split(",")

    for parts in subject_parts:
        parts = parts.strip()
        name, types = parts.split("(", 1)
        name = name.strip()
        types = types.strip(")")

        details = []
        type_entries = types.split("|")
        for type_entry in type_entries:
            type_entry = type_entry.strip()
            details.append(SubjectDetail(subj_type=type_entry))
        subject = Subject(name=name, details=details)
        subjects.append(subject)
    return subjects


def load_data_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet1 = workbook["Groups"]
    groups = []
    for row in sheet1.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        group_name = row[0]
        students_count = row[1]
        pi_text = row[2]
        subjects = parse_groups_subjects(pi_text)
        groups.append(Group(group_name, students_count, subjects))

    sheet2 = workbook["Teachers"]
    teachers = []
    for row in sheet2.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        teacher_name = row[0]
        subjects_list = row[1]
        hours = row[2]

        subjects = parse_teachers_subjects(subjects_list)
        teachers.append(Teacher(teacher_name, subjects, hours))

    sheet3 = workbook["Auditoriums"]
    auditoriums = []
    for row in sheet3.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        num = row[0]
        capacity = row[1]
        auditoriums.append(Auditorium(num, capacity))
    return groups, teachers, auditoriums


def export_schedule_to_excel(schedule_to_save, groups, auditoriums, filename="schedule.xlsx"):
    lesson_dicts = [
        {
            **lesson.to_dict(),
            "day_num": lesson.day.value,
        }
        for lesson in schedule_to_save.lessons
    ]
    df = pd.DataFrame(lesson_dicts)

    column_for_group_order = [
        "Group",
        "Day",
        "Lesson_num",
        "Subject",
        "Lesson_type",
        "Teacher",
        "Subgroup",
        "Auditorium",
    ]
    column_for_teacher_order = [
        "Teacher",
        "Day",
        "Lesson_num",
        "Subject",
        "Lesson_type",
        "Group",
        "Subgroup",
        "Auditorium",
    ]

    sorted_by_group = (
        df[column_for_group_order + ["day_num"]]
        .sort_values(by=["Group", "day_num", "Lesson_num"])
        .drop(columns=["day_num"])
    )

    sorted_by_teacher = (
        df[column_for_teacher_order + ["day_num"]]
        .sort_values(by=["Teacher", "day_num", "Lesson_num"])
        .drop(columns=["day_num"])
    )

    auditorium_data = defaultdict(lambda: defaultdict(list))

    for lesson in schedule_to_save.lessons:
        auditorium_data[lesson.auditorium][(lesson.day, lesson.lesson_num)].append(lesson)

    group_student_count = {}
    for group in groups:
        group_student_count[group.name] = group.students_count

    auditorium_records = []
    for auditorium_num, lessons_by_day in auditorium_data.items():
        for (day, lesson_num), lessons in lessons_by_day.items():
            groups = [lesson.group for lesson in lessons]
            total_students = sum(group_student_count[group] for group in groups)
            auditorium_capacity = next(a.capacity for a in auditoriums if a.number == auditorium_num)

            auditorium_records.append({
                "Auditorium": auditorium_num,
                "Day": Day(day).name,
                "Lesson_num": lesson_num,
                "Subject": lessons[0].subject,
                "Groups": ", ".join(set(groups)),
                "Subgroup": ", ".join(str(lesson.subgroup) for lesson in lessons if lesson.subgroup),
                "Total_students": total_students,
                "Capacity": auditorium_capacity,
            })

    sorted_by_auditorium = pd.DataFrame(auditorium_records)

    sorted_by_auditorium = sorted_by_auditorium.sort_values(
        by=["Auditorium", "Day", "Lesson_num"],
        ascending=[True, True, True]
    )

    with pd.ExcelWriter(filename) as writer:
        sorted_by_group.to_excel(writer, sheet_name="Sorted_By_Groups", index=False)
        sorted_by_teacher.to_excel(writer, sheet_name="Sorted_By_Teachers", index=False)
        sorted_by_auditorium.to_excel(writer, sheet_name="Sorted_By_Auditorium", index=False)

    print(f"Saved in '{filename}'")


def test_generate(group_q, teacher_q, aud_q, subj_q):
    subjects = []
    teachers = []
    groups = []
    auditoriums = []

    for i in range(subj_q):
        name = f"Subject{i + 1}"
        possible_hours = [21, 42, 63]

        details = [
            SubjectDetail(subj_type="Lec", hours=random.choice(possible_hours)),
        ]

        if random.random() < 0.75:
            details.append(
                SubjectDetail(
                    subj_type="Lab",
                    hours=random.choice(possible_hours),
                    subgroups=random.randint(1, 3),
                )
            )

        subjects.append(Subject(name=name, details=details))

    for i in range(teacher_q):
        name = f"Teacher{i + 1}"
        teacher_subjects = []

        assigned_subjects = random.sample(subjects, random.randint(2, min(12, len(subjects))))

        for subject in assigned_subjects:
            subject_details = []

            for detail in subject.details:
                if detail.type == "Lab":
                    subject_details.append(detail)

            if any(d.type == "Lec" for d in subject.details) and random.random() < 0.7:
                subject_details.extend([d for d in subject.details if d.type == "Lec"])

            if subject_details:
                teacher_subjects.append(
                    Subject(name=subject.name, details=subject_details)
                )

        hours = random.randint(20, 35)
        teachers.append(Teacher(name=name, subjects=teacher_subjects, hours=hours))

    for i in range(group_q):
        name = f"Group{i + 1}"
        students_count = random.randint(20, 40)

        assigned_subjects = random.sample(subjects, random.randint(6, 8))

        group_subjects = []
        for subject in assigned_subjects:
            subject_details = []

            subject_details.extend([d for d in subject.details if d.type == "Lec"])

            if any(d.type == "Lab" for d in subject.details) and random.random() < 0.75:
                subject_details.extend([d for d in subject.details if d.type == "Lab"])

            group_subjects.append(Subject(name=subject.name, details=subject_details))

        groups.append(
            Group(name=name, students_count=students_count, subjects=group_subjects)
        )

    for i in range(aud_q):
        number = f"A{i + 1}"
        capacity = random.randint(20, 120)
        auditoriums.append(Auditorium(number=number, capacity=capacity))
    return subjects, teachers, groups, auditoriums
