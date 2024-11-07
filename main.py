import random

import openpyxl

from classes import *


def parse_groups_subjects(pi_text):
    subjects = []
    subject_parts = pi_text.split("-")

    for part in subject_parts:
        if "(" in part and ")" in part:
            subject_name = part.split("(")[0].strip()
            details_text = part.split("(")[1].rstrip().rstrip(")")
            details = []
            for detail in details_text.split(","):
                parts = detail.split("|")
                type_ = parts[0].strip()
                hours = float(parts[1])
                subgroups = int(parts[2]) if len(parts) > 2 else None
                details.append(SubjectDetail(type_, hours, subgroups))
            subjects.append(Subject(subject_name, details))
    return subjects


def parse_teachers_subjects(text):
    subjects = []
    subject_parts = text.split(",")

    for parts in subject_parts:
        parts = parts.strip()
        name, types = parts.split("(", 1)
        name = name.strip()
        types = types.strip(")")

        details = []
        type_entries = types.split("|")
        for type_entry in type_entries:
            type_entry = type_entry.strip()
            details.append(SubjectDetail(subj_type=type_entry))
        subject = Subject(name=name, details=details)
        subjects.append(subject)
    return subjects


def load_data_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet1 = workbook["Groups"]
    groups = []
    for row in sheet1.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        group_name = row[0]
        students_count = row[1]
        pi_text = row[2]
        subjects = parse_groups_subjects(pi_text)
        groups.append(Group(group_name, students_count, subjects))

    sheet2 = workbook["Teachers"]
    teachers = []
    for row in sheet2.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        teacher_name = row[0]
        subjects_list = row[1]
        hours = row[2]

        subjects = parse_teachers_subjects(subjects_list)
        teachers.append(Teacher(teacher_name, subjects, hours))

    sheet3 = workbook["Auditoriums"]
    auditoriums = []
    for row in sheet3.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        num = row[0]
        capacity = row[1]
        auditoriums.append(Auditorium(num, capacity))
    return groups, teachers, auditoriums


def test_generate(group_q, teacher_q, aud_q, subj_q):
    subjects = []
    teachers = []
    groups = []
    auditoriums = []

    import random

    for i in range(subj_q):
        name = f"Subject{i + 1}"
        details = [
            SubjectDetail(subj_type="Lec", hours=round(random.uniform(10, 42) * 2) / 2)
        ]

        if random.random() < 0.75:
            details.append(
                SubjectDetail(
                    subj_type="Lab",
                    hours=round(random.uniform(10, 42) * 2) / 2,
                    subgroups=random.randint(1, 3),
                )
            )

        subjects.append(Subject(name=name, details=details))

    for i in range(teacher_q):
        name = f"Teacher{i + 1}"
        teacher_subjects = []

        assigned_subjects = random.sample(subjects, random.randint(4, 12))

        for subject in assigned_subjects:
            subject_details = []

            for detail in subject.details:
                if detail.type == "Lab":
                    subject_details.append(detail)

            if any(d.type == "Lec" for d in subject.details) and random.random() < 0.7:
                subject_details.extend([d for d in subject.details if d.type == "Lec"])

            if subject_details:
                teacher_subjects.append(
                    Subject(name=subject.name, details=subject_details)
                )

        hours = random.randint(10, 30)
        teachers.append(Teacher(name=name, subjects=teacher_subjects, hours=hours))

    for i in range(group_q):
        name = f"Group{i + 1}"
        students_count = random.randint(20, 40)

        assigned_subjects = random.sample(subjects, random.randint(6, 8))

        group_subjects = []
        for subject in assigned_subjects:
            subject_details = []

            subject_details.extend([d for d in subject.details if d.type == "Lec"])

            if any(d.type == "Lab" for d in subject.details) and random.random() < 0.75:
                subject_details.extend([d for d in subject.details if d.type == "Lab"])

            group_subjects.append(Subject(name=subject.name, details=subject_details))

        groups.append(
            Group(name=name, students_count=students_count, subjects=group_subjects)
        )

    for i in range(aud_q):
        number = f"A{i + 1}"
        capacity = random.randint(20, 120)
        auditoriums.append(Auditorium(number=number, capacity=capacity))
    return subjects, teachers, groups, auditoriums


def output_of_input(groups, teachers, auditoriums):
    for group in groups:
        print(f"Group: {group.name}, Student quantity: {group.students_count}")
        for subject in group.subjects:
            print(f"  Subject: {subject.name}")
            for detail in subject.details:
                print(
                    f"    Type: {detail.type}, Hours: {detail.hours}, Subgroups: {detail.subgroups}"
                )

    print("\nTeachers:")
    for teacher in teachers:
        print(f"Teacher: {teacher.name}, Hours: {teacher.hours}")
        for subject in teacher.subjects:
            print(f"  Subject: {subject.name}")
            for detail in subject.details:
                print(f"    Type: {detail.type}")

    print("\nAuditoriums:")
    for auditorium in auditoriums:
        print(f"auditorium №{auditorium.number}, Capacity: {auditorium.capacity}")


def fitness(schedule, groups, teachers, auditoriums, week_quantity):
    score = 0

    window_group_penalty = 0
    for group in groups:
        for day in Day:
            day_lessons = sorted(
                [lesson for lesson in schedule.lessons if
                 lesson.group == group.name and lesson.day == day],
                key=lambda x: x.lesson_num
            )
            for i in range(len(day_lessons) - 1):
                if (day_lessons[i + 1].lesson_num - day_lessons[i].lesson_num) > 1:
                    window_group_penalty += 1

    score -= window_group_penalty
    print(score)

    window_teacher_penalty = 0
    for teacher in teachers:
        for day in Day:
            day_lessons = sorted(
                [lesson for lesson in schedule.lessons if
                 lesson.teacher == teacher.name and lesson.day == day],
                key=lambda x: x.lesson_num
            )
            for i in range(len(day_lessons) - 1):
                if (day_lessons[i + 1].lesson_num - day_lessons[i].lesson_num) > 1:
                    window_teacher_penalty += 1

    score -= window_teacher_penalty
    print(score)

    capacity_penalty = 0
    for lesson in schedule.lessons:
        group = next(g for g in groups if g.name == lesson.group)
        auditorium = next(a for a in auditoriums if a.number == lesson.auditorium)
        if group.students_count > auditorium.capacity:
            capacity_penalty += 1

    score -= capacity_penalty
    print(score)

    weekly_hours_penalty = 0
    for teacher in teachers:
        total_hours = sum(
            1.5 for lesson in schedule.lessons if lesson.teacher == teacher.name
        )
        if total_hours > teacher.hours:
            weekly_hours_penalty += total_hours - teacher.hours

    score -= weekly_hours_penalty
    print(score)

    overlearning_time_penalty = 0
    for group in groups:
        subject_dict = {}
        for lesson in schedule.lessons:
            if lesson.group == group.name:
                name = str(lesson.subject + "-" + lesson.lesson_type)
                if not name in subject_dict:
                    subject_dict[name] = 1
                else:
                    subject_dict[name] += 1
        for subject_item in subject_dict:
            split_array = subject_item.split("-")
            subj_name = split_array[0]
            subj_type = split_array[1]
            result_detail = next(
                (detail for subject in group.subjects if subject.name == subj_name
                 for detail in subject.details if detail.type == subj_type),
                None
            )
            subgroups_count = result_detail.subgroups
            if not subgroups_count:
                subgroups_count = 1
            subject_time = abs(
                (subject_dict[subject_item] * 1.5 * week_quantity) / subgroups_count - result_detail.hours)
            overlearning_time_penalty += subject_time

    score -= (overlearning_time_penalty / week_quantity)
    print("Overlearning time:", (overlearning_time_penalty / week_quantity))
    print("Final score: ", score)

    return score


def crossover(schedule1, schedule2, groups, teachers, week_quantity, max_lessons_per_day=4):
    new_schedule = Schedule()

    for group in groups:
        assigned_teachers = {}

        for day in Day:
            for lesson_num in range(1, max_lessons_per_day + 1):
                parent_schedule = schedule1 if random.random() < 0.5 else schedule2
                lesson = next(
                    (l for l in parent_schedule.lessons
                     if l.group == group.name and l.day == day and l.lesson_num == lesson_num),
                    None
                )

                if lesson:
                    subject = lesson.subject
                    lesson_type = lesson.lesson_type

                    if (subject, lesson_type) not in assigned_teachers:
                        assigned_teachers[(subject, lesson_type)] = lesson.teacher

                    if lesson.teacher == assigned_teachers[(subject, lesson_type)]:
                        if new_schedule._check_constraints(lesson):
                            new_schedule.lessons.append(lesson)
                        else:
                            alt_schedule = schedule2 if parent_schedule == schedule1 else schedule1
                            alt_lesson = next(
                                (l for l in alt_schedule.lessons
                                 if l.group == group.name and l.day == day and l.lesson_num == lesson_num),
                                None
                            )
                            if (alt_lesson and
                                    alt_lesson.teacher == assigned_teachers[(subject, lesson_type)] and
                                    new_schedule._check_constraints(alt_lesson)):
                                new_schedule.lessons.append(alt_lesson)

    mutated_schedule = mutation_fixed_group_subjects(new_schedule, groups, teachers, week_quantity)
    smoothed_schedule = smoothing(mutated_schedule, teachers)
    return mutation_fixed_group_subjects(smoothed_schedule, groups, teachers, week_quantity)


def mutation_fixed_group_subjects(schedule, groups, teachers, week_quantity):
    for group in groups:
        for subject in group.subjects:
            for detail in subject.details:
                subgroups_count = 1
                if detail.subgroups is not None:
                    subgroups_count = detail.subgroups

                for subgroup in range(1, subgroups_count + 1):
                    subgroup_name = str(str(subgroup) + "/" + str(subgroups_count))
                    subject_hours = sum(
                        1.5 for lesson in schedule.lessons
                        if lesson.group == group.name and lesson.subject == subject.name and
                        detail.type == lesson.lesson_type and (detail.subgroups == 1 or detail.subgroups is None or
                                                               lesson.subgroup == subgroup_name)
                    )
                    max_hours = detail.hours / week_quantity

                    if subject_hours > max_hours:
                        excess_hours = subject_hours - max_hours
                        lessons_to_remove = [
                            lesson for lesson in schedule.lessons
                            if lesson.group == group.name and lesson.subject == subject.name and
                               detail.type == lesson.lesson_type and (
                                       detail.subgroups == 1 or detail.subgroups is None or
                                       lesson.subgroup == subgroup_name)
                        ]
                        for _ in range(int(excess_hours / 1.5)):
                            lesson_to_random = []
                            for lesson in lessons_to_remove:
                                if lesson.lesson_num == 1 or lesson.lesson_num == 4:
                                    lesson_to_random.append(lesson)
                            if len(lesson_to_random) < 1:
                                lesson_to_random = lessons_to_remove
                            to_delete = random.choice(lesson_to_random)
                            lessons_to_remove.remove(to_delete)
                            schedule.lessons.remove(to_delete)

                    if subject_hours < max_hours / 2:
                        lack_hours = max_hours - subject_hours
                        existing_lesson = [
                            lesson for lesson in schedule.lessons
                            if lesson.group == group.name and lesson.subject == subject.name and
                               detail.type == lesson.lesson_type and (
                                       detail.subgroups == 1 or detail.subgroups is None or
                                       lesson.subgroup == subgroup_name)
                        ]
                        available_teachers = []
                        if not existing_lesson:
                            available_teachers = [
                                t
                                for t in teachers
                                if any(
                                    ts.name == subject.name and max_hours <= t.hours
                                    and detail.type in [tsd.type for tsd in ts.details]
                                    for ts in t.subjects
                                )
                            ]
                        else:
                            available_teachers.append(existing_lesson[0].teacher)

                        for _ in range(int(lack_hours / 1.5)):
                            days_list = list(Day)
                            random.shuffle(days_list)
                            assigned = False
                            for day in days_list:
                                for i in range(2, 4):
                                    flag = 0
                                    for lesson in schedule.lessons:
                                        if (lesson.group == group.name and lesson.day == day and
                                                lesson.lesson_num == i and (
                                                        detail.subgroups == 1 or detail.subgroups is None or
                                                        lesson.subgroup == subgroup_name)):
                                            break
                                        if (lesson.group == group.name and lesson.day == day and
                                                lesson.lesson_num == i - 1):
                                            if subgroup is not None and subgroup != 1:
                                                if lesson.subgroup == subgroup_name:
                                                    flag += 1
                                            else:
                                                flag += 1
                                        if (lesson.group == group.name and lesson.day == day and
                                                lesson.lesson_num == i + 1):
                                            if subgroup is not None and subgroup != 1:
                                                if lesson.subgroup == subgroup_name:
                                                    flag += 1
                                            else:
                                                flag += 1
                                    if flag > 1:
                                        for teacher in available_teachers:
                                            auditorium = random.choice(
                                                [
                                                    a
                                                    for a in auditoriums
                                                    if a.capacity >= group.students_count
                                                ]
                                            )
                                            if not auditorium:
                                                auditorium = random.choice(
                                                    [
                                                        a
                                                        for a in auditoriums
                                                    ]
                                                )

                                            if detail.type == "Lec" or subgroups_count == 1:
                                                subgroup_name = None

                                            lesson = Lesson(
                                                day,
                                                i,
                                                teacher,
                                                detail.type,
                                                subject,
                                                group,
                                                auditorium,
                                                subgroup_name,
                                            )
                                            if schedule._check_constraints(lesson):
                                                schedule.lessons.append(lesson)
                                                assigned = True
                                                break
                            loop_num = 0
                            while not assigned and loop_num < 100:
                                teacher = random.choice(available_teachers)
                                day = random.choice(days_list)
                                lesson_num = random.randint(1, 4)
                                auditorium = random.choice(
                                    [
                                        a
                                        for a in auditoriums
                                        if a.capacity >= group.students_count
                                    ]
                                )
                                if not auditorium:
                                    auditorium = random.choice(
                                        [
                                            a
                                            for a in auditoriums
                                        ]
                                    )

                                if detail.type == "Lec" or subgroups_count == 1:
                                    subgroup_name = None

                                lesson = Lesson(
                                    day,
                                    lesson_num,
                                    teacher,
                                    detail.type,
                                    subject,
                                    group,
                                    auditorium,
                                    subgroup_name,
                                )
                                loop_num += 1
                                if schedule._check_constraints(lesson):
                                    schedule.lessons.append(lesson)
                                    assigned = True
    return schedule


def smoothing(new_schedule, teachers):
    for teacher in teachers:
        teacher_hours = sum(
            1.5 for lesson in new_schedule.lessons if lesson.teacher == teacher.name
        )

        if teacher_hours > teacher.hours:
            excess_hours = teacher_hours - teacher.hours

            lessons_by_subject_type = {}
            for lesson in new_schedule.lessons:
                if lesson.teacher == teacher.name:
                    key = (lesson.subject, lesson.lesson_type, lesson.group)
                    if key not in lessons_by_subject_type:
                        lessons_by_subject_type[key] = []
                    lessons_by_subject_type[key].append(lesson)

            while excess_hours > 0 and lessons_by_subject_type:
                random_key = random.choice(list(lessons_by_subject_type.keys()))
                random_group_lessons = lessons_by_subject_type[random_key]

                if excess_hours >= 1.5 * (len(random_group_lessons) - 1):
                    for lesson in random_group_lessons:
                        new_schedule.lessons.remove(lesson)
                    excess_hours -= 1.5 * len(random_group_lessons)

                del lessons_by_subject_type[random_key]

            if excess_hours <= 0:
                break

    return new_schedule


if __name__ == "__main__":
    file_path = "schedule_data.xlsx"
    init_groups, init_teachers, init_auditoriums = load_data_from_excel(file_path)
    gen_subjects, gen_teachers, gen_groups, gen_auditoriums = test_generate(
        # 28, 16, 8, 18
        0, 0, 0, 0
    )
    week_quantity = 14

    groups = init_groups + gen_groups
    teachers = init_teachers + gen_teachers
    auditoriums = init_auditoriums + gen_auditoriums

    schedules_collection = []


    # child_schedule = crossover(schedules_collection[0], schedules_collection[1], groups, teachers, week_quantity)
    # print(fitness(child_schedule, groups, teachers, auditoriums, week_quantity))
    # child_schedule.export_schedule_to_excel("schedule.xlsx")

    specimen_num = 100
    iter_num = 50

    for i in range(specimen_num):
        schedule = Schedule()
        schedule.generate_schedule(
            groups=groups, teachers=teachers, auditoriums=auditoriums, week_quantity=week_quantity
        )
        fitness_score = fitness(schedule, groups, teachers, auditoriums, week_quantity)
        # print(fitness_score)
        # schedule.export_schedule_to_excel(f"schedule{i}.xlsx")
        schedules_collection.append((schedule, fitness_score))

    for i in range(iter_num):
        sorted_items = sorted(
            schedules_collection,
            key=lambda x: x[1],
            reverse=True
        )
        half_size = len(sorted_items) // 2
        sorted_items = sorted_items[:half_size]
        schedules_collection = sorted_items

        print()
        print(f"Best iter {i}: ")
        for fi in range(0, len(sorted_items)):
            print(sorted_items[fi][1])

        for j in range(half_size):
            random_parent1 = random.randint(0, half_size - 1)
            random_parent2 = random.randint(0, half_size - 1)
            while random_parent1 == random_parent2:
                random_parent2 = random.randint(0, half_size - 1)

            child_schedule = crossover(sorted_items[random_parent1][0],
                                       sorted_items[random_parent2][0], groups, teachers, week_quantity)
            fitness_score = fitness(child_schedule, groups, teachers, auditoriums, week_quantity)
            schedules_collection.append((child_schedule, fitness_score))

    schedules_collection = sorted(
        schedules_collection,
        key=lambda x: x[1],
        reverse=True
    )

    # print()
    # print("BEST LAST ITER:")
    # for fi in range(0, len(schedules_collection)):
    #     print(schedules_collection[fi][1])

    print()
    print("BEST:")
    fitness(schedules_collection[0][0], groups, teachers, auditoriums, week_quantity)
    schedules_collection[0][0].export_schedule_to_excel(f"schedule.xlsx")

